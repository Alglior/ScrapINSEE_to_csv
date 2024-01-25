import requests
from bs4 import BeautifulSoup
import os
import pandas as pd
from io import StringIO
import re
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Function to add borders to all cells in a sheet
def add_borders_to_sheet(ws):
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

# Function to extract the code from the URL
def extract_code_from_url(url):
    match = re.search(r'COM-(\d+)', url)
    return match.group(1) if match else "unknown"

# URL of the web page you want to analyze
url = "https://www.insee.fr/fr/statistiques/2011101?geo=COM-42330"

# Send a GET request to fetch the page content
response = requests.get(url)

if response.status_code == 200:
    soup = BeautifulSoup(response.text, 'html.parser')
    tables = soup.find_all('table')
    
    output_directory = "tableaux_excel"
    if not os.path.exists(output_directory):
        os.mkdir(output_directory)
    
    sanitized_url = re.sub(r'[\\/*?:"<>|]', "", url.replace("https://", ""))
    url_directory = os.path.join(output_directory, sanitized_url)
    if not os.path.exists(url_directory):
        os.mkdir(url_directory)
    
    com_number = extract_code_from_url(url)

    mots_cles = ["POP", "FAM", "LOG", "FOR", "EMP", "ACT", "RFD", "REV", "DEN", "TOU"]

    for index, table in enumerate(tables):
        table_text = table.get_text().strip().upper()
        
        for mot_cle in mots_cles:
            if mot_cle in table_text:
                keyword_directory = os.path.join(url_directory, mot_cle)
                if not os.path.exists(keyword_directory):
                    os.mkdir(keyword_directory)
                
                h3_title = table.find_previous("h3")
                if h3_title:
                    title_text = re.sub(r'[\\/*?:"<>|\n\r]+', "", h3_title.text.strip())
                    title_text = (title_text[:50] + '..') if len(title_text) > 50 else title_text
                else:
                    title_text = f"Tableau_{index + 1}"
                
                html_str = str(table)
                # Replace commas with dots in the HTML content to preserve decimal points
                html_str = html_str.replace(',', '.')
                df = pd.read_html(StringIO(html_str))[0]
                
                if 'Unnamed: 0' in df.columns:
                    df.drop(columns=['Unnamed: 0'], inplace=True)
                
                # Add a new 'Code' column with the code extracted from the URL
                df['Code'] = com_number
                
                # Create an Excel workbook for each table
                wb = Workbook()
                ws = wb.active
                
                # Add DataFrame data to the Excel sheet
                for r_idx, df_row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(df_row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        
                        # Check if the cell value is numeric and has a dot (.) as decimal separator
                        if isinstance(value, (int, float)) and '.' in str(value):
                            # Apply a custom number format to preserve dots as decimal separators
                            cell.number_format = '0.00'
                
                # Add borders to the cells
                add_borders_to_sheet(ws)

                # Construct the Excel file name
                excel_filename = os.path.join(keyword_directory, f"{com_number}_{title_text}.xlsx")
                wb.save(excel_filename)

    print("Les tableaux ont été classés et convertis en fichiers Excel avec des bordures autour des cellules.")
else:
    print(f"La requête a échoué avec le code d'état {response.status_code}")
