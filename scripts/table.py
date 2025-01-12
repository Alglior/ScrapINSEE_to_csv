import requests
from bs4 import BeautifulSoup
import os
import pandas as pd
from io import StringIO
import re
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import argparse

def add_borders_to_sheet(ws):
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

def extract_code_from_url(url):
    match = re.search(r'COM-(\d+)', url)
    return match.group(1) if match else "unknown"

def parse_arguments():
    parser = argparse.ArgumentParser()
    parser.add_argument('--progress-file', help='File to track download progress')
    parser.add_argument('--urls-file', required=True, help='File containing URLs to process')
    parser.add_argument('--resume', action='store_true', help='Resume from last successful download')
    return parser.parse_args()

def main():
    args = parse_arguments()
    
    # Initialiser le compteur de progression
    current_progress = 0
    
    # Si on reprend le téléchargement, lire la dernière progression
    if args.resume and args.progress_file and os.path.exists(args.progress_file):
        try:
            with open(args.progress_file, 'r') as f:
                current_progress = int(f.read().strip() or 0)
        except:
            current_progress = 0
    
    # Lire les URLs
    with open(args.urls_file, 'r') as f:
        urls = f.readlines()
    
    # Filtrer les URLs non vides et les nettoyer
    urls = [url.strip() for url in urls if url.strip()]
    
    # Si on reprend, ignorer les URLs déjà traitées
    if current_progress > 0:
        urls = urls[current_progress:]

    # Modifier le chemin du dossier de sortie pour qu'il soit au même niveau que le dossier scripts
    output_directory = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "tableaux_excel")
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    downloaded_count = current_progress
    for url in urls:
        response = requests.get(url)

        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            tables = soup.find_all('table')
            
            sanitized_url = re.sub(r'[\\/*?:"<>|]', "", url.replace("https://", ""))
            url_directory = os.path.join(output_directory, sanitized_url)
            if not os.path.exists(url_directory):
                os.mkdir(url_directory)
            
            com_number = extract_code_from_url(url)

            for index, table in enumerate(tables):
                html_str = str(table)
                html_str = html_str.replace(',', '.')
                df = pd.read_html(StringIO(html_str))[0]
                df.columns = ['_'.join(map(str, col)).strip().replace('U_n_n_a_m_e_d_:_ _0', '').replace(' ', '_') for col in df.columns.values]
                df['Code'] = com_number

                wb = Workbook()
                ws = wb.active

                for r_idx, df_row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(df_row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        if isinstance(value, (int, float)) and '.' in str(value):
                            cell.number_format = '0.00'

                add_borders_to_sheet(ws)

                h3_title = table.find_previous("h3")
                title_text = f"Tableau_{index + 1}" if not h3_title else re.sub(r'[\\/*?:"<>|\n\r]+', "", h3_title.text.strip())[:100]
                
                excel_filename = os.path.join(url_directory, f"{com_number}_{title_text}.xlsx")
                wb.save(excel_filename)

            print(f"Tous les tableaux de l'URL '{url}' ont été enregistrés dans le dossier '{url_directory}'.")
        else:
            print(f"La requête pour l'URL '{url}' a échoué avec le code d'état {response.status_code}")

        downloaded_count += 1
        if args.progress_file:
            try:
                with open(args.progress_file, 'w') as f:
                    f.write(str(downloaded_count))
            except Exception:
                pass

    # À la fin du script, assurez-vous d'écrire le nombre total
    if args.progress_file:
        try:
            with open(args.progress_file, 'w') as f:
                f.write(str(downloaded_count))
        except Exception:
            pass

if __name__ == "__main__":
    main()
